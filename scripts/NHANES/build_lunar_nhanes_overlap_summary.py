#!/usr/bin/env python3
"""
Build LUNAR_NHANES_OVERLAP_Summary.xlsx from nhanes_biopro_all_cycles_combined.csv.

Purpose
-------
Creates a NHANES-only overlap workbook containing only chemistry variables that
are expected to overlap with Inspiration4 and/or bed-rest clinical chemistry
summaries. The workbook is intentionally cross-sectional: SEQN is the subject ID
and no timepoint-specific sheets are produced.

Input
-----
processed_csv/nhanes_biopro_all_cycles_combined.csv

Outputs
-------
summaries/LUNAR_NHANES_OVERLAP_Summary_v2.xlsx
processed_csv/LUNAR_NHANES_OVERLAP_Data_Wide_v2.csv

Notes
-----
- Includes 16 direct BIOPRO chemistry variables.
- Adds 2 derived variables: Albumin/Globulin Ratio and BUN/Creatinine Ratio.
- The Excel workbook includes an Overlap_Data preview; the full overlap-ready
  participant-level dataset is exported as CSV.
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DIRECT_OVERLAP_META: dict[str, dict[str, str]] = {
    "LBXSAL": {"canonical": "Albumin", "unit": "g/dL", "description": "Albumin, refrigerated serum"},
    "LBXSAPSI": {"canonical": "Alkaline Phosphatase", "unit": "U/L", "description": "Alkaline Phosphatase (ALP)"},
    "LBXSASSI": {"canonical": "AST", "unit": "U/L", "description": "Aspartate Aminotransferase (AST)"},
    "LBXSATSI": {"canonical": "ALT", "unit": "U/L", "description": "Alanine Aminotransferase (ALT)"},
    "LBXSBU": {"canonical": "BUN", "unit": "mg/dL", "description": "Blood Urea Nitrogen (BUN)"},
    "LBXSC3SI": {"canonical": "Carbon Dioxide / Bicarbonate", "unit": "mmol/L", "description": "Bicarbonate / Carbon Dioxide"},
    "LBXSCA": {"canonical": "Calcium", "unit": "mg/dL", "description": "Total Calcium"},
    "LBXSCLSI": {"canonical": "Chloride", "unit": "mmol/L", "description": "Chloride"},
    "LBXSCR": {"canonical": "Creatinine", "unit": "mg/dL", "description": "Creatinine, refrigerated serum"},
    "LBXSGB": {"canonical": "Globulin", "unit": "g/dL", "description": "Globulin, calculated"},
    "LBXSGL": {"canonical": "Glucose", "unit": "mg/dL", "description": "Glucose, refrigerated serum"},
    "LBXSKSI": {"canonical": "Potassium", "unit": "mmol/L", "description": "Potassium"},
    "LBXSNASI": {"canonical": "Sodium", "unit": "mmol/L", "description": "Sodium"},
    "LBXSTB": {"canonical": "Total Bilirubin", "unit": "mg/dL", "description": "Total Bilirubin"},
    "LBXSTP": {"canonical": "Total Protein", "unit": "g/dL", "description": "Total Protein"},
    "LBXSUA": {"canonical": "Uric Acid", "unit": "mg/dL", "description": "Uric Acid"},
}

DERIVED_DEFINITIONS: dict[str, dict[str, str]] = {
    "ALBUMIN_GLOBULIN_RATIO": {
        "canonical": "Albumin/Globulin Ratio",
        "unit": "ratio",
        "description": "Albumin/Globulin Ratio",
        "formula": "LBXSAL / LBXSGB",
        "inputs": "LBXSAL, LBXSGB",
    },
    "BUN_CREATININE_RATIO": {
        "canonical": "BUN/Creatinine Ratio",
        "unit": "ratio",
        "description": "BUN/Creatinine Ratio",
        "formula": "LBXSBU / LBXSCR",
        "inputs": "LBXSBU, LBXSCR",
    },
}


def safe_stats(series: pd.Series) -> dict[str, Any]:
    s = pd.to_numeric(series, errors="coerce")
    valid = s.dropna()
    if valid.empty:
        return {"N": 0, "Mean": None, "SD": None, "Median": None, "Min": None, "Max": None,
                "Missing_Count": int(s.isna().sum()), "Missing_Percent": float(s.isna().mean() * 100)}
    return {"N": int(valid.count()), "Mean": float(valid.mean()),
            "SD": float(valid.std(ddof=1)) if valid.count() > 1 else 0.0,
            "Median": float(valid.median()), "Min": float(valid.min()), "Max": float(valid.max()),
            "Missing_Count": int(s.isna().sum()), "Missing_Percent": float(s.isna().mean() * 100)}


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    bottom_border = Border(bottom=Side(style="thin", color="D9D9D9"))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.border = bottom_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    for idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 42)


def add_table(ws) -> None:
    if ws.max_row <= 1 or ws.max_column <= 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    name = "T_" + "".join(ch for ch in ws.title if ch.isalnum())[:20]
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def write_rows(wb: Workbook, title: str, rows: list[list[Any]], table: bool = True):
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    # Full styling over tens of thousands of rows is slow. For large sheets,
    # style only the header and set freeze panes/widths.
    if len(rows) > 10000:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        white_font = Font(color="FFFFFF", bold=True)
        bottom_border = Border(bottom=Side(style="thin", color="D9D9D9"))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.border = bottom_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for idx, col_cells in enumerate(ws.iter_cols(max_row=min(ws.max_row, 2000)), start=1):
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 42)
    else:
        style_sheet(ws)
        if table:
            add_table(ws)
    return ws


def build(input_csv: Path, output_xlsx: Path, output_csv: Path, id_col: str = "SEQN", preview_rows: int = 5000) -> None:
    df = pd.read_csv(input_csv)
    if id_col not in df.columns:
        raise ValueError(f"Expected ID column {id_col!r} not found.")

    direct_vars = [v for v in DIRECT_OVERLAP_META if v in df.columns]
    overlap = df[[id_col] + direct_vars].copy()

    # Derived variables are generated only when all required inputs exist.
    derived_meta: dict[str, dict[str, str]] = {}
    if {"LBXSAL", "LBXSGB"}.issubset(overlap.columns):
        overlap["ALBUMIN_GLOBULIN_RATIO"] = np.where(
            overlap["LBXSGB"].notna() & (overlap["LBXSGB"] != 0),
            overlap["LBXSAL"] / overlap["LBXSGB"],
            np.nan,
        )
        derived_meta["ALBUMIN_GLOBULIN_RATIO"] = DERIVED_DEFINITIONS["ALBUMIN_GLOBULIN_RATIO"]
    if {"LBXSBU", "LBXSCR"}.issubset(overlap.columns):
        overlap["BUN_CREATININE_RATIO"] = np.where(
            overlap["LBXSCR"].notna() & (overlap["LBXSCR"] != 0),
            overlap["LBXSBU"] / overlap["LBXSCR"],
            np.nan,
        )
        derived_meta["BUN_CREATININE_RATIO"] = DERIVED_DEFINITIONS["BUN_CREATININE_RATIO"]

    all_meta = {**DIRECT_OVERLAP_META, **derived_meta}
    overlap_vars = [c for c in overlap.columns if c != id_col]

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    overlap.to_csv(output_csv, index=False)

    variable_stats = []
    isv_rows = []
    missing_rows = []
    for var in overlap_vars:
        meta = all_meta.get(var, {})
        st = safe_stats(overlap[var])
        variable_stats.append([var, meta.get("canonical", var), st["N"], st["Mean"], st["SD"], st["Median"], st["Min"], st["Max"], st["Missing_Count"], st["Missing_Percent"]])
        cv = None if st["Mean"] in (None, 0) else (st["SD"] / st["Mean"] * 100)
        isv_rows.append([var, meta.get("canonical", var), st["N"], st["Mean"], st["SD"], cv, st["Median"], st["Min"], st["Max"]])
        missing_rows.append([var, meta.get("canonical", var), st["Missing_Count"], st["Missing_Percent"]])

    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset_Inventory"
    inventory = [
        ["Metric", "Value"],
        ["Dataset", "NHANES BIOPRO all-cycles combined"],
        ["Workbook Purpose", "NHANES-only overlap summary for LUNAR harmonization"],
        ["ID Column", id_col],
        ["N_Rows", int(len(overlap))],
        ["N_Unique_Subjects", int(overlap[id_col].nunique())],
        ["Duplicate_Subject_IDs", int(len(overlap) - overlap[id_col].nunique())],
        ["N_Overlap_Variables_Excluding_ID", int(len(overlap_vars))],
        ["N_Direct_Overlap_Variables", int(len(direct_vars))],
        ["N_Derived_Overlap_Variables", int(len(derived_meta))],
        ["Generated", dt.datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Notes", "SEQN is treated strictly as participant ID and excluded from all variable-level statistics."],
    ]
    for row in inventory:
        ws.append(row)
    style_sheet(ws)

    write_rows(wb, "Variable_Stats", [["Variable", "Canonical_Name", "N", "Mean", "SD", "Median", "Min", "Max", "Missing_Count", "Missing_Percent"]] + variable_stats)
    write_rows(wb, "Inter_Subject_Variability", [["Variable", "Canonical_Name", "N_Subjects", "Mean_of_Subject_Means", "Between_Subject_SD", "CV_Percent", "Median_of_Subject_Means", "Min_Subject_Mean", "Max_Subject_Mean"]] + isv_rows)
    write_rows(wb, "Missing_By_Variable", [["Variable", "Canonical_Name", "Missing_Count", "Missing_Percent"]] + missing_rows)

    p_missing = overlap[overlap_vars].isna().sum(axis=1)
    p_missing_pct = p_missing / len(overlap_vars) * 100
    part_rows = [[id_col, "Missing_Count", "Missing_Percent"]]
    for sid, mc, mp in zip(overlap[id_col], p_missing, p_missing_pct):
        part_rows.append([sid, int(mc), float(mp)])
    write_rows(wb, "Missing_By_Participant", part_rows, table=False)

    units_rows = [["Variable", "Canonical_Name", "Unit", "Variable_Type"]]
    naming_rows = [["Variable", "Canonical_Name", "Description", "Variable_Type"]]
    for var in overlap_vars:
        meta = all_meta.get(var, {})
        variable_type = "Derived" if var in derived_meta else "Direct"
        units_rows.append([var, meta.get("canonical", var), meta.get("unit", ""), variable_type])
        naming_rows.append([var, meta.get("canonical", var), meta.get("description", ""), variable_type])
    write_rows(wb, "Units_Audit", units_rows)
    write_rows(wb, "Naming_Audit", naming_rows)

    derived_rows = [["Variable", "Canonical_Name", "Formula", "Inputs", "Unit", "Notes"]]
    for var, meta in derived_meta.items():
        derived_rows.append([var, meta["canonical"], meta["formula"], meta["inputs"], meta["unit"], "Derived from NHANES BIOPRO variables."])
    write_rows(wb, "Derived_Variables_Audit", derived_rows)

    preview = overlap.head(preview_rows).replace({np.nan: None})
    overlap_data_rows = [list(preview.columns)] + preview.values.tolist()
    write_rows(wb, "Overlap_Data", overlap_data_rows, table=False)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build the LUNAR NHANES overlap summary workbook.")
    parser.add_argument("--input", default="processed_csv/nhanes_biopro_all_cycles_combined.csv", type=Path)
    parser.add_argument("--output-xlsx", default="summaries/LUNAR_NHANES_OVERLAP_Summary_v2.xlsx", type=Path)
    parser.add_argument("--output-csv", default="processed_csv/LUNAR_NHANES_OVERLAP_Data_Wide_v2.csv", type=Path)
    parser.add_argument("--id-col", default="SEQN")
    parser.add_argument("--preview-rows", default=5000, type=int)
    args = parser.parse_args()
    build(args.input, args.output_xlsx, args.output_csv, args.id_col, args.preview_rows)
