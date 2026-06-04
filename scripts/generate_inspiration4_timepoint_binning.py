#!/usr/bin/env python3
"""
generate_inspiration4_timepoint_binning.py

Purpose
-------
Generate an Inspiration4 timepoint-binning workbook from the LUNAR Inspiration4
overlap workbook.

The script creates:
1. One worksheet per variable with:
   - raw values by subject and timepoint
   - Option 1: mean preflight vs mean postflight
   - Option 2: acute response, L-3 vs R+1
   - Option 3: long-term recovery, mean preflight vs postflight excluding R+1

2. Three cohort-level summary worksheets:
   - Summary_Option1
   - Summary_Option2
   - Summary_Option3

Expected input columns
----------------------
The input workbook should contain a sheet named "Overlap_Data" with at least:
- Subject_ID
- Variable
- Timepoint
- Value

Usage
-----
python generate_inspiration4_timepoint_binning.py \
    --input LUNAR_Inspiration4_OVERLAP_Summary_v1.xlsx \
    --output Inspiration4_Timepoint_Binning_Workbook_All3_Summaries.xlsx
"""

import argparse
import re
import statistics
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


TIMEPOINT_ORDER = ["L-92", "L-44", "L-3", "R+1", "R+45", "R+82", "R+194"]

OPTION_DEFINITIONS = {
    "Option 1: Mean Preflight vs Mean Postflight": {
        "baseline_label": "Preflight_Mean",
        "post_label": "Postflight_Mean",
        "baseline_timepoints": ["L-92", "L-44", "L-3"],
        "post_timepoints": ["R+1", "R+45", "R+82", "R+194"],
    },
    "Option 2: Acute Response (L-3 vs R+1)": {
        "baseline_label": "L-3",
        "post_label": "R+1",
        "baseline_timepoints": ["L-3"],
        "post_timepoints": ["R+1"],
    },
    "Option 3: Long-term Recovery": {
        "baseline_label": "Preflight_Mean",
        "post_label": "Postflight_Mean_No_R+1",
        "baseline_timepoints": ["L-92", "L-44", "L-3"],
        "post_timepoints": ["R+45", "R+82", "R+194"],
    },
}


def clean_sheet_title(title: str, existing_titles: set[str]) -> str:
    """Create an Excel-safe sheet title, limited to 31 characters."""
    cleaned = re.sub(r"[\[\]\*:/\\\?]", "_", str(title)).strip()
    cleaned = cleaned[:31] or "Variable"

    base = cleaned
    counter = 1
    while cleaned in existing_titles:
        suffix = f"_{counter}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1

    existing_titles.add(cleaned)
    return cleaned


def mean_or_blank(values):
    """Mean of non-missing values; returns None if all values are missing."""
    vals = [v for v in values if pd.notna(v)]
    if not vals:
        return None
    return float(pd.Series(vals).mean())


def mean_sd(values):
    """Return sample mean and sample SD for numeric values."""
    vals = [v for v in values if isinstance(v, (int, float)) and pd.notna(v)]
    if not vals:
        return None, None
    if len(vals) == 1:
        return float(vals[0]), 0
    return float(statistics.mean(vals)), float(statistics.stdev(vals))


def style_header(row):
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)


def write_variable_sheet(ws, variable_name, variable_df):
    """Write raw timepoint values and all three option calculations for one variable."""
    pivot = variable_df.pivot_table(
        index="Subject_ID",
        columns="Timepoint",
        values="Value",
        aggfunc="first",
    )

    pivot = pivot.reindex(columns=[tp for tp in TIMEPOINT_ORDER if tp in pivot.columns])
    pivot = pivot.sort_index()

    # Raw timepoint section
    ws.cell(1, 1, f"Variable: {variable_name}")
    ws.cell(1, 1).font = Font(bold=True, size=12)

    raw_header_row = 3
    headers = ["Subject_ID"] + list(pivot.columns)

    for c, header in enumerate(headers, start=1):
        ws.cell(raw_header_row, c, header)

    style_header(ws[raw_header_row])

    for r_idx, (subject_id, row) in enumerate(pivot.iterrows(), start=raw_header_row + 1):
        ws.cell(r_idx, 1, subject_id)
        for c_idx, value in enumerate(row, start=2):
            if pd.notna(value):
                ws.cell(r_idx, c_idx, float(value))

    # Option calculation sections
    current_row = raw_header_row + len(pivot) + 4

    for option_title, option in OPTION_DEFINITIONS.items():
        ws.cell(current_row, 1, option_title)
        ws.cell(current_row, 1).font = Font(bold=True, size=11)

        headers = [
            "Subject_ID",
            option["baseline_label"],
            option["post_label"],
            "Delta",
            "Percent_Delta",
            "N_Baseline",
            "N_Postflight",
        ]

        header_row = current_row + 1
        for c, header in enumerate(headers, start=1):
            ws.cell(header_row, c, header)

        style_header(ws[header_row])

        for r_idx, (subject_id, row) in enumerate(pivot.iterrows(), start=header_row + 1):
            baseline_values = row.reindex(option["baseline_timepoints"])
            post_values = row.reindex(option["post_timepoints"])

            baseline = mean_or_blank(baseline_values)
            post = mean_or_blank(post_values)
            delta = None if baseline is None or post is None else post - baseline
            percent_delta = None
            if baseline not in (None, 0) and post is not None:
                percent_delta = (post - baseline) / baseline * 100

            ws.cell(r_idx, 1, subject_id)
            ws.cell(r_idx, 2, baseline)
            ws.cell(r_idx, 3, post)
            ws.cell(r_idx, 4, delta)
            ws.cell(r_idx, 5, percent_delta)
            ws.cell(r_idx, 6, int(baseline_values.notna().sum()))
            ws.cell(r_idx, 7, int(post_values.notna().sum()))

            if percent_delta is not None:
                ws.cell(r_idx, 5).number_format = "0.0"

        current_row = header_row + len(pivot) + 4

    autosize_columns(ws)


def collect_option_rows(ws, marker):
    """Extract subject-level option values from a variable sheet."""
    marker_row = None

    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == marker:
            marker_row = r
            break

    if marker_row is None:
        return [], [], [], [], []

    data_start = marker_row + 2
    baseline_vals, post_vals, delta_vals, percent_delta_vals = [], [], [], []
    subjects = []

    r = data_start
    while r <= ws.max_row:
        subject = ws.cell(r, 1).value
        if subject is None:
            break

        subjects.append(subject)
        baseline_vals.append(ws.cell(r, 2).value)
        post_vals.append(ws.cell(r, 3).value)
        delta_vals.append(ws.cell(r, 4).value)
        percent_delta_vals.append(ws.cell(r, 5).value)
        r += 1

    return subjects, baseline_vals, post_vals, delta_vals, percent_delta_vals


def write_summary_sheet(wb, sheet_name, option_title, baseline_label, post_label):
    ws = wb.create_sheet(sheet_name, 0)

    headers = [
        "Variable",
        f"{baseline_label}_Mean",
        f"{baseline_label}_SD",
        f"{post_label}_Mean",
        f"{post_label}_SD",
        "Delta_Mean",
        "Delta_SD",
        "Percent_Delta_Mean",
        "Percent_Delta_SD",
        "N_Subjects",
    ]

    for c, header in enumerate(headers, start=1):
        ws.cell(1, c, header)

    style_header(ws[1])

    out_row = 2
    summary_sheet_names = {"Summary_Option1", "Summary_Option2", "Summary_Option3"}

    for variable_sheet_name in wb.sheetnames:
        if variable_sheet_name in summary_sheet_names:
            continue

        variable_ws = wb[variable_sheet_name]
        subjects, baseline_vals, post_vals, delta_vals, percent_delta_vals = collect_option_rows(
            variable_ws, option_title
        )

        if not subjects:
            continue

        baseline_mean, baseline_sd = mean_sd(baseline_vals)
        post_mean, post_sd = mean_sd(post_vals)
        delta_mean, delta_sd = mean_sd(delta_vals)
        pct_mean, pct_sd = mean_sd(percent_delta_vals)

        ws.cell(out_row, 1, variable_sheet_name)
        ws.cell(out_row, 2, baseline_mean)
        ws.cell(out_row, 3, baseline_sd)
        ws.cell(out_row, 4, post_mean)
        ws.cell(out_row, 5, post_sd)
        ws.cell(out_row, 6, delta_mean)
        ws.cell(out_row, 7, delta_sd)
        ws.cell(out_row, 8, pct_mean)
        ws.cell(out_row, 9, pct_sd)
        ws.cell(out_row, 10, len(subjects))

        ws.cell(out_row, 8).number_format = "0.0"
        ws.cell(out_row, 9).number_format = "0.0"

        out_row += 1

    autosize_columns(ws)


def generate_workbook(input_path, output_path):
    df = pd.read_excel(input_path, sheet_name="Overlap_Data")

    required_columns = {"Subject_ID", "Variable", "Timepoint", "Value"}
    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"Input file is missing required columns: {sorted(missing)}")

    df = df.copy()
    df["Variable"] = df["Variable"].astype(str)
    df["Timepoint"] = df["Timepoint"].astype(str)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    existing_titles = set()
    variable_name_to_sheet = {}

    for variable in sorted(df["Variable"].dropna().unique()):
        sheet_title = clean_sheet_title(variable, existing_titles)
        variable_name_to_sheet[variable] = sheet_title

        ws = wb.create_sheet(sheet_title)
        variable_df = df[df["Variable"] == variable]
        write_variable_sheet(ws, variable, variable_df)

    # Add summary sheets in reverse order because each is inserted at index 0.
    write_summary_sheet(
        wb,
        "Summary_Option3",
        "Option 3: Long-term Recovery",
        "Preflight",
        "Postflight_No_R1",
    )
    write_summary_sheet(
        wb,
        "Summary_Option2",
        "Option 2: Acute Response (L-3 vs R+1)",
        "L3",
        "R1",
    )
    write_summary_sheet(
        wb,
        "Summary_Option1",
        "Option 1: Mean Preflight vs Mean Postflight",
        "Preflight",
        "Postflight",
    )

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Generate Inspiration4 binned timepoint workbook with all three cohort summaries."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to input workbook, e.g., LUNAR_Inspiration4_OVERLAP_Summary_v1.xlsx",
    )
    parser.add_argument(
        "--output",
        default="Inspiration4_Timepoint_Binning_Workbook_All3_Summaries.xlsx",
        help="Path to output workbook.",
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    generate_workbook(input_path, output_path)
    print(f"Saved workbook to: {output_path}")


if __name__ == "__main__":
    main()
