#!/usr/bin/env python3
"""
generate_bedrest_timepoint_binning.py

Purpose
-------
Generate a Bed Rest Campaign 1 timepoint-binning workbook from the LUNAR Bed Rest
overlap workbook.

The script creates:
1. One worksheet per overlap variable with:
   - subject-level phase means for Pre_BedRest, In_BedRest, and Post_BedRest
   - subject-level post-pre delta

2. One cohort-level summary worksheet:
   - Cohort_Summary

Expected input columns
----------------------
The input workbook should contain a sheet named "Overlap_Data" with at least:
- Subject_ID
- Test_Phase
- Value_Numeric
- Inspiration4_Variable

The script uses Inspiration4_Variable as the harmonized overlap variable name so that
the output is directly comparable to the Inspiration4 binned workbook.

Usage
-----
python generate_bedrest_timepoint_binning.py \
    --input LUNAR_BedRest_Campaign1_OVERLAP_Summary_v3.xlsx \
    --output BedRest_Timepoint_Binning_Workbook.xlsx
"""

import argparse
import re
import statistics
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


PHASE_MAP = {
    "PRE_TEST": "Pre_BedRest",
    "IN_TEST": "In_BedRest",
    "POST_TEST": "Post_BedRest",
}

PHASE_ORDER = ["Pre_BedRest", "In_BedRest", "Post_BedRest"]


def clean_sheet_title(title: str, existing_titles: set[str]) -> str:
    """Create an Excel-safe worksheet title, limited to 31 characters."""
    cleaned = re.sub(r"[\[\]\*:/\\\?]", "_", str(title)).strip()
    cleaned = cleaned[:31] or "Variable"

    base = cleaned
    counter = 1
    while cleaned in existing_titles:
        suffix = f"_{counter}"
        cleaned = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1

    existing_titles.add(cleaned)
    return cleaned


def mean_sd(values):
    """Return sample mean and sample SD for numeric values."""
    vals = [v for v in values if isinstance(v, (int, float)) and pd.notna(v)]

    if not vals:
        return None, None
    if len(vals) == 1:
        return float(vals[0]), 0

    return float(statistics.mean(vals)), float(statistics.stdev(vals))


def style_header(row):
    """Apply consistent header styling."""
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws):
    """Autosize worksheet columns for readability."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)


def write_variable_sheet(ws, variable_name, variable_df):
    """
    Write one worksheet for a harmonized overlap variable.

    Values are averaged within each bed rest phase for each subject:
    - PRE_TEST  -> Pre_BedRest
    - IN_TEST   -> In_BedRest
    - POST_TEST -> Post_BedRest
    """
    pivot = variable_df.pivot_table(
        index="Subject_ID",
        columns="Bin",
        values="Value_Numeric",
        aggfunc="mean",
    )

    pivot = pivot.reindex(columns=[phase for phase in PHASE_ORDER if phase in pivot.columns])
    pivot = pivot.sort_index()

    ws.cell(1, 1, f"Variable: {variable_name}")
    ws.cell(1, 1).font = Font(bold=True, size=12)

    # Phase means section
    raw_header_row = 3
    headers = ["Subject_ID"] + list(pivot.columns)

    for c, header in enumerate(headers, start=1):
        ws.cell(raw_header_row, c, header)

    style_header(ws[raw_header_row])

    for r_idx, (subject_id, row) in enumerate(pivot.iterrows(), start=raw_header_row + 1):
        ws.cell(r_idx, 1, subject_id)

        for c_idx, value in enumerate(row, start=2):
            if pd.notna(value):
                ws.cell(r_idx, c_idx, float(value))

    # Subject-level summary section
    summary_start = raw_header_row + len(pivot) + 4

    ws.cell(summary_start, 1, "Subject-Level Summary")
    ws.cell(summary_start, 1).font = Font(bold=True, size=11)

    summary_headers = [
        "Subject_ID",
        "Pre_BedRest",
        "In_BedRest",
        "Post_BedRest",
        "Post-Pre Delta",
        "Percent_Delta",
    ]

    for c, header in enumerate(summary_headers, start=1):
        ws.cell(summary_start + 1, c, header)

    style_header(ws[summary_start + 1])

    pre_vals = []
    in_vals = []
    post_vals = []
    delta_vals = []
    pct_delta_vals = []

    for r_idx, (subject_id, row) in enumerate(pivot.iterrows(), start=summary_start + 2):
        pre = row.get("Pre_BedRest")
        in_bedrest = row.get("In_BedRest")
        post = row.get("Post_BedRest")

        delta = None
        pct_delta = None

        if pd.notna(pre) and pd.notna(post):
            delta = float(post - pre)
            if pre != 0:
                pct_delta = float((post - pre) / pre * 100)

        ws.cell(r_idx, 1, subject_id)

        if pd.notna(pre):
            ws.cell(r_idx, 2, float(pre))
            pre_vals.append(float(pre))

        if pd.notna(in_bedrest):
            ws.cell(r_idx, 3, float(in_bedrest))
            in_vals.append(float(in_bedrest))

        if pd.notna(post):
            ws.cell(r_idx, 4, float(post))
            post_vals.append(float(post))

        if delta is not None:
            ws.cell(r_idx, 5, delta)
            delta_vals.append(delta)

        if pct_delta is not None:
            ws.cell(r_idx, 6, pct_delta)
            ws.cell(r_idx, 6).number_format = "0.0"
            pct_delta_vals.append(pct_delta)

    autosize_columns(ws)

    return {
        "Variable": variable_name,
        "Pre_Mean": mean_sd(pre_vals)[0],
        "Pre_SD": mean_sd(pre_vals)[1],
        "In_Mean": mean_sd(in_vals)[0],
        "In_SD": mean_sd(in_vals)[1],
        "Post_Mean": mean_sd(post_vals)[0],
        "Post_SD": mean_sd(post_vals)[1],
        "Delta_Mean": mean_sd(delta_vals)[0],
        "Delta_SD": mean_sd(delta_vals)[1],
        "Percent_Delta_Mean": mean_sd(pct_delta_vals)[0],
        "Percent_Delta_SD": mean_sd(pct_delta_vals)[1],
        "N_Subjects_Pre": len(pre_vals),
        "N_Subjects_In": len(in_vals),
        "N_Subjects_Post": len(post_vals),
        "N_Subjects_Delta": len(delta_vals),
    }


def write_cohort_summary(wb, summary_rows):
    """Write cohort-level summary sheet."""
    ws = wb.create_sheet("Cohort_Summary", 0)

    headers = [
        "Variable",
        "Pre_Mean",
        "Pre_SD",
        "In_Mean",
        "In_SD",
        "Post_Mean",
        "Post_SD",
        "Delta_Mean",
        "Delta_SD",
        "Percent_Delta_Mean",
        "Percent_Delta_SD",
        "N_Subjects_Pre",
        "N_Subjects_In",
        "N_Subjects_Post",
        "N_Subjects_Delta",
    ]

    for c, header in enumerate(headers, start=1):
        ws.cell(1, c, header)

    style_header(ws[1])

    for r_idx, row in enumerate(summary_rows, start=2):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx, row.get(header))

        ws.cell(r_idx, 10).number_format = "0.0"
        ws.cell(r_idx, 11).number_format = "0.0"

    autosize_columns(ws)


def generate_workbook(input_path, output_path):
    """Generate the full Bed Rest timepoint-binning workbook."""
    df = pd.read_excel(input_path, sheet_name="Overlap_Data")

    required_columns = {
        "Subject_ID",
        "Test_Phase",
        "Value_Numeric",
        "Inspiration4_Variable",
    }

    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"Input file is missing required columns: {sorted(missing)}")

    df = df.copy()
    df = df[df["Value_Numeric"].notna()].copy()

    df["Test_Phase"] = df["Test_Phase"].astype(str)
    df["Bin"] = df["Test_Phase"].map(PHASE_MAP)

    # Keep only rows that map cleanly into the three bed rest phases.
    df = df[df["Bin"].notna()].copy()

    df["Inspiration4_Variable"] = df["Inspiration4_Variable"].astype(str)

    wb = Workbook()
    wb.remove(wb.active)

    existing_titles = set()
    summary_rows = []

    for variable in sorted(df["Inspiration4_Variable"].dropna().unique()):
        sheet_title = clean_sheet_title(variable, existing_titles)
        ws = wb.create_sheet(sheet_title)

        variable_df = df[df["Inspiration4_Variable"] == variable].copy()
        summary_row = write_variable_sheet(ws, variable, variable_df)
        summary_rows.append(summary_row)

    write_cohort_summary(wb, summary_rows)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Generate a Bed Rest Campaign 1 binned phase workbook with cohort summary."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to input workbook, e.g., LUNAR_BedRest_Campaign1_OVERLAP_Summary_v3.xlsx",
    )
    parser.add_argument(
        "--output",
        default="BedRest_Timepoint_Binning_Workbook.xlsx",
        help="Path to output workbook.",
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    generate_workbook(input_path, output_path)
    print(f"Saved workbook to: {output_path}")


if __name__ == "__main__":
    main()
