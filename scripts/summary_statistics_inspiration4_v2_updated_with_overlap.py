#!/usr/bin/env python3
"""
Build Inspiration4 data-quality / summary-statistics workbooks used in LUNAR.

This script intentionally does NOT summarize the original raw assay files directly.
It reproduces the cleaned Inspiration4 summary workbook from the already-harmonized
master files:

    1. Inspiration4_Master_Wide.csv
    2. Inspiration4_Master_Long.csv
    3. Demographics_inspiration4.xlsx

It can also generate a second Inspiration4-to-Bed Rest overlap summary workbook
using a locked crosswalk of canonical overlap variables. This overlap workbook is
designed to mirror the Bed Rest Campaign 1 overlap workbook and includes only the
approved overlap variables.

Example: full Inspiration4 summary
----------------------------------
python scripts/summary_statistics_inspiration4_v2.py \
    --wide outputs/Inspiration4_Master_Wide.csv \
    --long outputs/Inspiration4_Master_Long.csv \
    --demographics data/raw/Inspiration4/Demographics_inspiration4.xlsx \
    --out outputs/Inspiration4_Data_Quality_Summary_V2.xlsx

Example: full summary + overlap summary
---------------------------------------
python scripts/summary_statistics_inspiration4_v2.py \
    --wide outputs/Inspiration4_Master_Wide.csv \
    --long outputs/Inspiration4_Master_Long.csv \
    --demographics data/raw/Inspiration4/Demographics_inspiration4.xlsx \
    --out outputs/Inspiration4_Data_Quality_Summary_V2.xlsx \
    --overlap-out outputs/Inspiration4_BedRest_Overlap_Summary_v3_with_BUN.xlsx

Outputs
-------
1. Inspiration4_Data_Quality_Summary_V2.xlsx
   Full Inspiration4 QC workbook built from the master wide/long files.

2. Inspiration4_BedRest_Overlap_Summary_v3_with_BUN.xlsx
   Optional overlap-only workbook using the locked Inspiration4-to-Bed Rest
   variable crosswalk. This includes BUN and UREA NITROGEN (BUN) as separate
   canonical labels mapped to the same Inspiration4 source column.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_BASELINE_TIMEPOINTS = ["L-92", "L-44", "L-3"]

# Bed Rest overlap summary tab set that the Inspiration4 workbook is modeled after.
BEDREST_ALIGNED_TABS = [
    "Dataset_Inventory",
    "Variable_Stats",
    "Variable_Stats_By_Timepoint",
    "Subject_Stats",
    "Subject_Timepoint_Variable_Stat",
    "Subject_Missingness",
    "Subject_Timepoint_Stats",
    "Missing_By_Variable",
    "Missing_By_Timepoint",
    "Overlap_Crosswalk",
    "Units_Audit",
    "Derived_Variables_Audit",
    "Participant_Audit",
    "Overlap_Data",
]

EXTRA_INSPIRATION4_TABS = [
    "Demographics",
    "Missing_By_Participant",
    "Missing_By_Part_Time",
    "Naming_Audit",
    "Inter_Subject_Variability",
    "Baseline_ISV",
    "Variable_Coverage",
    "Long_File_Variable_Inventory",
]

# Locked Inspiration4-to-Bed Rest overlap crosswalk.
#
# IMPORTANT:
# This is intentionally hard-coded rather than fuzzy-matched to avoid accidentally
# including similarly named variables. The overlap workbook should contain only
# these canonical variables.
#
# BUN and UREA NITROGEN (BUN) are both retained as canonical labels for
# harmonization purposes, and both map to the same Inspiration4 source column.
LOCKED_BEDREST_OVERLAP_CROSSWALK = {
    "ALBUMIN": "CMP__ALBUMIN",
    "ALKALINE PHOSPHATASE": "CMP__ALKALINE PHOSPHATASE",
    "ALT": "CMP__ALT",
    "AST": "CMP__AST",
    "BILIRUBIN; TOTAL": "CMP__BILIRUBIN; TOTAL",
    "CALCIUM": "CMP__CALCIUM",
    "CARBON DIOXIDE": "CMP__CARBON DIOXIDE",
    "CHLORIDE": "CMP__CHLORIDE",
    "CREATININE": "CMP__CREATININE",
    "CRP": "Cardiovascular_Eve__CRP",
    "GLUCOSE": "CMP__GLUCOSE",
    "POTASSIUM": "CMP__POTASSIUM",
    "PROTEIN; TOTAL": "CMP__PROTEIN; TOTAL",
    "SODIUM": "CMP__SODIUM",
    "BUN": "CMP__UREA NITROGEN (BUN)",
    "UREA NITROGEN (BUN)": "CMP__UREA NITROGEN (BUN)",
    "GLOBULIN": "CMP__GLOBULIN",
    "ALBUMIN/GLOBULIN RATIO": "CMP__ALBUMIN/GLOBULIN RATIO",
    "BUN/CREATININE RATIO": "CMP__BUN/CREATININE RATIO",
}


# -----------------------------------------------------------------------------
# I/O helpers
# -----------------------------------------------------------------------------

def read_csv_flex(path: Path) -> pd.DataFrame:
    """Read a CSV using common encodings."""
    errors = []
    for enc in ("utf-8", "utf-8-sig", "latin1", "cp1252"):
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception as exc:  # noqa: BLE001 - preserve attempts for final error
            errors.append(f"{enc}: {exc}")
    raise ValueError(f"Could not read CSV {path}. Attempts: {' | '.join(errors)}")


def read_demographics(path: Path | None) -> pd.DataFrame:
    """Read the first non-empty worksheet from the demographics workbook."""
    if path is None or not path.exists():
        return pd.DataFrame(
            {"Note": ["No demographics workbook provided. Demographics were not available."]}
        )

    xls = pd.ExcelFile(path)
    for sheet in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        if not df.empty and df.shape[1] > 0:
            df = df.loc[:, ~df.columns.astype(str).str.contains(r"^Unnamed", regex=True)]
            return df

    return pd.DataFrame({"Note": ["No non-empty sheet found in demographics workbook."]})


def clean_sheet_name(name: str) -> str:
    """Excel-safe sheet name, max 31 chars."""
    return re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name))[:31]


def find_col(df: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    """Find a column by exact or partial case-insensitive name match."""
    lower_to_original = {c.lower().strip(): c for c in df.columns}
    for candidate in candidates:
        key = candidate.lower().strip()
        if key in lower_to_original:
            return lower_to_original[key]

    for col in df.columns:
        col_lower = col.lower().strip()
        for candidate in candidates:
            if candidate.lower().strip() in col_lower:
                return col
    return None


def split_feature(feature_id: str) -> tuple[str, str]:
    """Split a wide-master feature column into panel and variable."""
    if "__" in str(feature_id):
        panel, variable = str(feature_id).split("__", 1)
        return panel, variable
    return "", str(feature_id)


def q1(x: pd.Series) -> float:
    return x.quantile(0.25)


def q3(x: pd.Series) -> float:
    return x.quantile(0.75)


# -----------------------------------------------------------------------------
# Workbook writer
# -----------------------------------------------------------------------------

def write_df(ws, df: pd.DataFrame) -> None:
    """Write dataframe values to a worksheet."""
    df = df.replace([np.inf, -np.inf], np.nan)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        values = []
        for value in row:
            if pd.isna(value):
                values.append(None)
            elif isinstance(value, np.integer):
                values.append(int(value))
            elif isinstance(value, np.floating):
                values.append(float(value))
            else:
                values.append(value)
        ws.append(values)


def style_sheet(ws) -> None:
    """Apply simple, consistent workbook styling."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_gray = Side(style="thin", color="D9E1F2")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter][: min(ws.max_row, 250)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 55)


def save_workbook(
    tables: dict[str, pd.DataFrame],
    out_path: Path,
    ordered_names: list[str] | None = None,
) -> None:
    """Save all summary tables to a styled Excel workbook and validate it reopens."""
    if ordered_names is None:
        ordered_names = list(tables)

    ordered_names = [name for idx, name in enumerate(ordered_names) if name not in ordered_names[:idx]]
    ordered_names.extend([name for name in tables if name not in ordered_names])

    wb = Workbook()
    wb.remove(wb.active)

    for name in ordered_names:
        if name not in tables:
            continue
        ws = wb.create_sheet(clean_sheet_name(name))
        write_df(ws, tables[name])
        style_sheet(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    test = load_workbook(out_path, read_only=True, data_only=True)
    test.close()


# -----------------------------------------------------------------------------
# Full Inspiration4 summary table builders
# -----------------------------------------------------------------------------

def build_internal_long(wide: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, str, str, list[str]]:
    """Convert the master wide file into an internal long table for calculations."""
    subject_col = find_col(
        wide,
        ["Subject_ID", "Subject", "Participant", "Participant_ID", "Crew", "Crew_ID"],
    )
    timepoint_col = find_col(
        wide,
        ["Timepoint", "Time Point", "Collection_Timepoint", "Visit", "Mission_Day"],
    )

    if subject_col is None or timepoint_col is None:
        raise ValueError(
            "Could not identify subject/timepoint columns in the wide master. "
            f"Columns found: {wide.columns.tolist()}"
        )

    feature_cols = [c for c in wide.columns if c not in [subject_col, timepoint_col]]
    wide_numeric = wide.copy()
    for col in feature_cols:
        wide_numeric[col] = pd.to_numeric(wide_numeric[col], errors="coerce")

    feature_meta = pd.DataFrame(
        [(feature, *split_feature(feature)) for feature in feature_cols],
        columns=["Feature_ID", "Panel", "Variable"],
    )

    long = (
        wide_numeric.melt(
            id_vars=[subject_col, timepoint_col],
            value_vars=feature_cols,
            var_name="Feature_ID",
            value_name="Value",
        )
        .merge(feature_meta, on="Feature_ID", how="left")
    )

    return long, feature_meta, subject_col, timepoint_col, feature_cols


def build_summary_tables(
    wide: pd.DataFrame,
    long_inventory: pd.DataFrame,
    demographics: pd.DataFrame,
    baseline_timepoints: list[str],
) -> dict[str, pd.DataFrame]:
    """Build all full Inspiration4 workbook tabs from master files."""
    long, feature_meta, subject_col, timepoint_col, feature_cols = build_internal_long(wide)

    subjects = sorted(wide[subject_col].dropna().astype(str).unique())
    timepoints = sorted(wide[timepoint_col].dropna().astype(str).unique())
    n_subjects = len(subjects)
    n_timepoints = len(timepoints)
    n_features = len(feature_cols)
    expected_total = n_subjects * n_timepoints * n_features
    present_total = int(long["Value"].notna().sum())
    missing_total = expected_total - present_total

    dataset_inventory = pd.DataFrame(
        [
            ["Dataset", "Inspiration4"],
            ["Source files used", "Inspiration4_Master_Wide.csv; Inspiration4_Master_Long.csv; Demographics_inspiration4.xlsx"],
            ["Rows in wide master", len(wide)],
            ["Subjects", n_subjects],
            ["Subject IDs", ", ".join(subjects)],
            ["Timepoints", n_timepoints],
            ["Timepoint labels", ", ".join(timepoints)],
            ["Feature columns in wide master", n_features],
            ["Unique base variable names", feature_meta["Variable"].nunique()],
            ["Expected subject-timepoint-feature cells", expected_total],
            ["Observed non-missing values", present_total],
            ["Missing cells", missing_total],
            ["Overall missing percent", round(missing_total / expected_total * 100, 4) if expected_total else np.nan],
            ["Raw source-file rows included?", "No"],
            ["Notes", "Summary generated from harmonized master wide/long files only; raw assay rows are not included."],
        ],
        columns=["Metric", "Value"],
    )

    variable_stats = (
        long.groupby(["Feature_ID", "Panel", "Variable"], dropna=False)["Value"]
        .agg(N="count", Mean="mean", SD="std", Median="median", Q1=q1, Q3=q3, Min="min", Max="max")
        .reset_index()
    )
    variable_stats["Expected_N"] = n_subjects * n_timepoints
    variable_stats["Missing_N"] = variable_stats["Expected_N"] - variable_stats["N"]
    variable_stats["Missing_Percent"] = 100 * variable_stats["Missing_N"] / variable_stats["Expected_N"]
    variable_stats["IQR"] = variable_stats["Q3"] - variable_stats["Q1"]
    variable_stats = variable_stats[
        [
            "Feature_ID", "Panel", "Variable", "N", "Expected_N", "Missing_N", "Missing_Percent",
            "Mean", "SD", "Median", "Q1", "Q3", "IQR", "Min", "Max",
        ]
    ]

    variable_stats_by_timepoint = (
        long.groupby(["Feature_ID", "Panel", "Variable", timepoint_col], dropna=False)["Value"]
        .agg(N="count", Mean="mean", SD="std", Median="median", Min="min", Max="max")
        .reset_index()
        .rename(columns={timepoint_col: "Timepoint"})
    )
    variable_stats_by_timepoint["Expected_N"] = n_subjects
    variable_stats_by_timepoint["Missing_N"] = variable_stats_by_timepoint["Expected_N"] - variable_stats_by_timepoint["N"]
    variable_stats_by_timepoint["Missing_Percent"] = (
        100 * variable_stats_by_timepoint["Missing_N"] / variable_stats_by_timepoint["Expected_N"]
    )

    subject_stats = (
        long.groupby([subject_col, "Feature_ID", "Panel", "Variable"], dropna=False)["Value"]
        .agg(N="count", Mean="mean", SD="std", Median="median", Min="min", Max="max")
        .reset_index()
        .rename(columns={subject_col: "Subject"})
    )
    subject_stats = subject_stats[
        ["Subject", "Feature_ID", "Panel", "Variable", "N", "Mean", "SD", "Median", "Min", "Max"]
    ]

    subject_timepoint_variable_stat = (
        long.groupby([subject_col, timepoint_col, "Feature_ID", "Panel", "Variable"], dropna=False)["Value"]
        .agg(N="count", Mean="mean", SD="std", Median="median", Min="min", Max="max")
        .reset_index()
        .rename(columns={subject_col: "Subject", timepoint_col: "Timepoint"})
    )

    missing_by_variable = variable_stats[
        ["Feature_ID", "Panel", "Variable", "Expected_N", "N", "Missing_N", "Missing_Percent"]
    ].rename(columns={"N": "Present_N"})

    missing_by_participant = (
        long.groupby(subject_col)["Value"]
        .agg(Present_N=lambda s: int(s.notna().sum()), Missing_N=lambda s: int(s.isna().sum()))
        .reset_index()
        .rename(columns={subject_col: "Subject"})
    )
    missing_by_participant["Expected_N"] = n_timepoints * n_features
    missing_by_participant["Missing_Percent"] = (
        100 * missing_by_participant["Missing_N"] / missing_by_participant["Expected_N"]
    )
    missing_by_participant = missing_by_participant[
        ["Subject", "Expected_N", "Present_N", "Missing_N", "Missing_Percent"]
    ]

    missing_by_timepoint = (
        long.groupby(timepoint_col)["Value"]
        .agg(Present_N=lambda s: int(s.notna().sum()), Missing_N=lambda s: int(s.isna().sum()))
        .reset_index()
        .rename(columns={timepoint_col: "Timepoint"})
    )
    missing_by_timepoint["Expected_N"] = n_subjects * n_features
    missing_by_timepoint["Missing_Percent"] = 100 * missing_by_timepoint["Missing_N"] / missing_by_timepoint["Expected_N"]
    missing_by_timepoint = missing_by_timepoint[
        ["Timepoint", "Expected_N", "Present_N", "Missing_N", "Missing_Percent"]
    ]

    missing_by_part_time = (
        long.groupby([subject_col, timepoint_col])["Value"]
        .agg(Present_N=lambda s: int(s.notna().sum()), Missing_N=lambda s: int(s.isna().sum()))
        .reset_index()
        .rename(columns={subject_col: "Subject", timepoint_col: "Timepoint"})
    )
    missing_by_part_time["Expected_N"] = n_features
    missing_by_part_time["Missing_Percent"] = 100 * missing_by_part_time["Missing_N"] / missing_by_part_time["Expected_N"]
    missing_by_part_time = missing_by_part_time[
        ["Subject", "Timepoint", "Expected_N", "Present_N", "Missing_N", "Missing_Percent"]
    ]

    subject_timepoint_stats = missing_by_part_time.rename(
        columns={"Present_N": "Variables_Present", "Missing_N": "Variables_Missing"}
    )

    participant_audit = (
        wide.groupby(subject_col)
        .agg(Records=(timepoint_col, "count"), Timepoints=(timepoint_col, "nunique"))
        .reset_index()
        .rename(columns={subject_col: "Subject"})
    )
    participant_audit["Expected_Feature_Cells"] = participant_audit["Records"] * n_features
    participant_audit["Observed_Feature_Values"] = participant_audit["Subject"].map(
        long.groupby(subject_col)["Value"].apply(lambda s: int(s.notna().sum()))
    )
    participant_audit["Missing_Feature_Values"] = (
        participant_audit["Expected_Feature_Cells"] - participant_audit["Observed_Feature_Values"]
    )
    participant_audit["Missing_Percent"] = (
        100 * participant_audit["Missing_Feature_Values"] / participant_audit["Expected_Feature_Cells"]
    )

    units_audit = feature_meta.copy()
    units_audit["Units"] = "Not provided in attached master wide/long files"
    units_audit = units_audit[["Feature_ID", "Panel", "Variable", "Units"]]

    naming_audit = feature_meta.copy()
    naming_audit["Duplicated_Base_Name_Across_Panels"] = naming_audit["Variable"].duplicated(keep=False)
    naming_audit = naming_audit[["Feature_ID", "Panel", "Variable", "Duplicated_Base_Name_Across_Panels"]]

    subject_means = (
        long.groupby(["Feature_ID", "Panel", "Variable", subject_col], dropna=False)["Value"]
        .mean()
        .reset_index()
        .rename(columns={subject_col: "Subject", "Value": "Subject_Mean"})
    )
    inter_subject_variability = (
        subject_means.groupby(["Feature_ID", "Panel", "Variable"], dropna=False)["Subject_Mean"]
        .agg(
            N_Subjects="count",
            Mean_of_Subject_Means="mean",
            Between_Subject_SD="std",
            Median_of_Subject_Means="median",
            Min_Subject_Mean="min",
            Max_Subject_Mean="max",
        )
        .reset_index()
    )
    inter_subject_variability["CV_Percent"] = (
        100 * inter_subject_variability["Between_Subject_SD"] / inter_subject_variability["Mean_of_Subject_Means"]
    )
    inter_subject_variability = inter_subject_variability[
        [
            "Feature_ID", "Panel", "Variable", "N_Subjects", "Mean_of_Subject_Means",
            "Between_Subject_SD", "CV_Percent", "Median_of_Subject_Means",
            "Min_Subject_Mean", "Max_Subject_Mean",
        ]
    ]

    baseline = long[long[timepoint_col].isin(baseline_timepoints)].copy()
    baseline_subject_means = (
        baseline.groupby(["Feature_ID", "Panel", "Variable", subject_col], dropna=False)["Value"]
        .mean()
        .reset_index()
        .rename(columns={subject_col: "Subject", "Value": "Baseline_Subject_Mean"})
    )
    baseline_isv = (
        baseline_subject_means.groupby(["Feature_ID", "Panel", "Variable"], dropna=False)["Baseline_Subject_Mean"]
        .agg(
            N_Subjects="count",
            Baseline_Mean_of_Subject_Means="mean",
            Baseline_Between_Subject_SD="std",
            Baseline_Median_of_Subject_Means="median",
            Baseline_Min_Subject_Mean="min",
            Baseline_Max_Subject_Mean="max",
        )
        .reset_index()
    )
    baseline_isv["Baseline_CV_Percent"] = (
        100 * baseline_isv["Baseline_Between_Subject_SD"] / baseline_isv["Baseline_Mean_of_Subject_Means"]
    )
    baseline_isv["Baseline_Timepoints"] = ", ".join(baseline_timepoints)
    baseline_isv = baseline_isv[
        [
            "Feature_ID", "Panel", "Variable", "Baseline_Timepoints", "N_Subjects",
            "Baseline_Mean_of_Subject_Means", "Baseline_Between_Subject_SD",
            "Baseline_CV_Percent", "Baseline_Median_of_Subject_Means",
            "Baseline_Min_Subject_Mean", "Baseline_Max_Subject_Mean",
        ]
    ]

    coverage_rows = []
    for (feature_id, panel, variable), group in long.groupby(["Feature_ID", "Panel", "Variable"], dropna=False):
        observed = group[group["Value"].notna()]
        expected_n = n_subjects * n_timepoints
        coverage_rows.append(
            {
                "Feature_ID": feature_id,
                "Panel": panel,
                "Variable": variable,
                "Expected_N": expected_n,
                "Observed_N": len(observed),
                "Missing_N": expected_n - len(observed),
                "Percent_Complete": 100 * len(observed) / expected_n,
                "Subjects_With_Data": observed[subject_col].nunique(),
                "Timepoints_With_Data": observed[timepoint_col].nunique(),
            }
        )
    variable_coverage = pd.DataFrame(coverage_rows)

    overlap_crosswalk = pd.DataFrame(
        {
            "Note": [
                "Not applicable for full-source Inspiration4 summary. Use the optional overlap workbook for Bed Rest overlap crosswalk."
            ]
        }
    )

    derived_variables_audit = pd.DataFrame(
        [
            ["GLOBULIN", "Not derived here; retained if present as a feature in the master wide file."],
            ["ALBUMIN/GLOBULIN RATIO", "Not derived here; retained if present as a feature in the master wide file."],
            ["BUN/CREATININE RATIO", "Not derived here; retained if present as a feature in the master wide file."],
        ],
        columns=["Variable", "Treatment"],
    )

    overlap_data = pd.DataFrame(
        {
            "Note": [
                "Raw observation-level overlap data are not included in this full summary workbook. "
                "Use the optional overlap workbook for Inspiration4/Bed Rest overlap data."
            ]
        }
    )

    if long_inventory.empty:
        long_file_variable_inventory = pd.DataFrame({"Note": ["No long/inventory file provided."]})
    elif long_inventory.shape[1] == 1:
        long_file_variable_inventory = long_inventory.copy()
        long_file_variable_inventory.columns = ["Variable_From_Attached_Long_File"]
    else:
        long_file_variable_inventory = long_inventory.copy()

    notes = pd.DataFrame(
        [
            ["Purpose", "Full Inspiration4 summary workbook aligned to the Bed Rest overlap summary structure."],
            ["Source files", "Master wide + master long/inventory + demographics workbook."],
            ["Raw data policy", "No raw source-file rows or raw observation-level export tabs are included."],
            ["Feature_ID", "Wide-file column, usually Panel__Variable; this prevents accidental cross-panel aggregation."],
            ["Baseline ISV", f"Calculated using {', '.join(baseline_timepoints)}."],
        ],
        columns=["Item", "Description"],
    )

    return {
        "v3_Notes": notes,
        "Dataset_Inventory": dataset_inventory,
        "Variable_Stats": variable_stats,
        "Variable_Stats_By_Timepoint": variable_stats_by_timepoint,
        "Subject_Stats": subject_stats,
        "Subject_Timepoint_Variable_Stat": subject_timepoint_variable_stat,
        "Subject_Missingness": missing_by_participant.copy(),
        "Subject_Timepoint_Stats": subject_timepoint_stats,
        "Missing_By_Variable": missing_by_variable,
        "Missing_By_Timepoint": missing_by_timepoint,
        "Overlap_Crosswalk": overlap_crosswalk,
        "Units_Audit": units_audit,
        "Derived_Variables_Audit": derived_variables_audit,
        "Participant_Audit": participant_audit,
        "Overlap_Data": overlap_data,
        "Demographics": demographics,
        "Missing_By_Participant": missing_by_participant,
        "Missing_By_Part_Time": missing_by_part_time,
        "Naming_Audit": naming_audit,
        "Inter_Subject_Variability": inter_subject_variability,
        "Baseline_ISV": baseline_isv,
        "Variable_Coverage": variable_coverage,
        "Long_File_Variable_Inventory": long_file_variable_inventory,
    }


# -----------------------------------------------------------------------------
# Overlap-only summary table builders
# -----------------------------------------------------------------------------

def build_overlap_long(wide: pd.DataFrame) -> tuple[pd.DataFrame, str, str]:
    """Build the locked Inspiration4/Bed Rest overlap long table."""
    subject_col = find_col(
        wide,
        ["Subject_ID", "Subject", "Participant", "Participant_ID", "Crew", "Crew_ID"],
    )
    timepoint_col = find_col(
        wide,
        ["Timepoint", "Time Point", "Collection_Timepoint", "Visit", "Mission_Day"],
    )

    if subject_col is None or timepoint_col is None:
        raise ValueError("Could not identify subject/timepoint columns in the wide master.")

    missing_columns = [
        (canonical, source_col)
        for canonical, source_col in LOCKED_BEDREST_OVERLAP_CROSSWALK.items()
        if source_col not in wide.columns
    ]
    if missing_columns:
        message = "\n".join([f"{canonical} -> {source}" for canonical, source in missing_columns])
        raise ValueError(
            "The following locked overlap source columns are missing from the wide master:\n"
            f"{message}"
        )

    records = []
    for _, row in wide.iterrows():
        for canonical, source_col in LOCKED_BEDREST_OVERLAP_CROSSWALK.items():
            records.append(
                {
                    "Subject_ID": row[subject_col],
                    "Timepoint": row[timepoint_col],
                    "Variable": canonical,
                    "Value": pd.to_numeric(row[source_col], errors="coerce"),
                    "Source_Column": source_col,
                }
            )

    return pd.DataFrame(records), subject_col, timepoint_col


def aggregate_overlap_stats(long: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    """Summary statistics for the overlap workbook."""
    return (
        long.groupby(group_cols, dropna=False)["Value"]
        .agg(N="count", Mean="mean", SD="std", Median="median", Min="min", Max="max")
        .reset_index()
    )


def overlap_missing_table(long: pd.DataFrame, group_cols: list[str], expected_per_group: int) -> pd.DataFrame:
    """Missingness table for the overlap workbook."""
    out = (
        long.groupby(group_cols, dropna=False)["Value"]
        .agg(Present_N=lambda x: int(x.notna().sum()), Missing_N=lambda x: int(x.isna().sum()))
        .reset_index()
    )
    out["Expected_N"] = expected_per_group
    out["Missing_Percent"] = 100 * out["Missing_N"] / out["Expected_N"]
    return out


def build_overlap_summary_tables(wide: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Build the Inspiration4/Bed Rest overlap-only summary workbook.

    This workbook uses the locked crosswalk and contains exactly the canonical
    variables defined in LOCKED_BEDREST_OVERLAP_CROSSWALK.
    """
    long, _, _ = build_overlap_long(wide)

    n_subjects = long["Subject_ID"].nunique()
    n_timepoints = long["Timepoint"].nunique()
    n_variables = long["Variable"].nunique()

    dataset_inventory = pd.DataFrame(
        [
            ["Dataset", "Inspiration4"],
            ["Source file", "Inspiration4_Master_Wide.csv"],
            ["Participants", n_subjects],
            ["Timepoints", n_timepoints],
            ["Canonical overlap variables", n_variables],
            ["Rows in Overlap_Data", len(long)],
            ["Crosswalk method", "Locked hard-coded mapping; no fuzzy matching"],
            [
                "Note",
                "BUN and UREA NITROGEN (BUN) are retained as separate canonical labels "
                "mapped to CMP__UREA NITROGEN (BUN).",
            ],
        ],
        columns=["Metric", "Value"],
    )

    subject_timepoint_stats = (
        long.groupby(["Subject_ID", "Timepoint"], dropna=False)["Value"]
        .agg(
            Variables_Present=lambda x: int(x.notna().sum()),
            Variables_Missing=lambda x: int(x.isna().sum()),
        )
        .reset_index()
    )
    subject_timepoint_stats["Variables_Expected"] = n_variables
    subject_timepoint_stats["Missing_Percent"] = (
        100 * subject_timepoint_stats["Variables_Missing"] / n_variables
    )

    derived_variables_audit = pd.DataFrame(
        [
            ["GLOBULIN", "Present in Inspiration4 master data", "CMP__GLOBULIN"],
            ["ALBUMIN/GLOBULIN RATIO", "Present in Inspiration4 master data", "CMP__ALBUMIN/GLOBULIN RATIO"],
            ["BUN/CREATININE RATIO", "Present in Inspiration4 master data", "CMP__BUN/CREATININE RATIO"],
            ["BUN", "Canonical alias retained for cross-dataset harmonization", "CMP__UREA NITROGEN (BUN)"],
        ],
        columns=["Variable", "Status", "Source_Column"],
    )

    return {
        "Dataset_Inventory": dataset_inventory,
        "Variable_Stats": aggregate_overlap_stats(long, ["Variable"]),
        "Variable_Stats_By_Timepoint": aggregate_overlap_stats(long, ["Variable", "Timepoint"]),
        "Subject_Stats": aggregate_overlap_stats(long, ["Subject_ID", "Variable"]),
        "Subject_Timepoint_Variable_Stat": aggregate_overlap_stats(
            long,
            ["Subject_ID", "Timepoint", "Variable"],
        ),
        "Subject_Missingness": overlap_missing_table(
            long,
            ["Subject_ID"],
            n_timepoints * n_variables,
        ),
        "Subject_Timepoint_Stats": subject_timepoint_stats,
        "Missing_By_Variable": overlap_missing_table(
            long,
            ["Variable"],
            n_subjects * n_timepoints,
        ),
        "Missing_By_Timepoint": overlap_missing_table(
            long,
            ["Timepoint"],
            n_subjects * n_variables,
        ),
        "Overlap_Crosswalk": pd.DataFrame(
            [
                {"Canonical_Variable": canonical, "Inspiration4_Column": source_col}
                for canonical, source_col in LOCKED_BEDREST_OVERLAP_CROSSWALK.items()
            ]
        ),
        "Units_Audit": pd.DataFrame(
            {
                "Variable": list(LOCKED_BEDREST_OVERLAP_CROSSWALK.keys()),
                "Source_Column": list(LOCKED_BEDREST_OVERLAP_CROSSWALK.values()),
                "Units": "See source CMP/Eve panel metadata",
            }
        ),
        "Derived_Variables_Audit": derived_variables_audit,
        "Participant_Audit": (
            long.groupby("Subject_ID", dropna=False)
            .agg(
                Records=("Value", "size"),
                Variables=("Variable", "nunique"),
                Timepoints=("Timepoint", "nunique"),
                Nonmissing_Values=("Value", lambda x: int(x.notna().sum())),
                Missing_Values=("Value", lambda x: int(x.isna().sum())),
            )
            .reset_index()
        ),
        "Overlap_Data": long[["Subject_ID", "Timepoint", "Variable", "Value", "Source_Column"]],
    }


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build Inspiration4 full and optional Bed Rest overlap summary workbooks from master files."
    )
    parser.add_argument("--wide", type=Path, required=True, help="Path to Inspiration4_Master_Wide.csv")
    parser.add_argument("--long", type=Path, required=True, help="Path to Inspiration4_Master_Long.csv or variable inventory CSV")
    parser.add_argument("--demographics", type=Path, default=None, help="Path to Demographics_inspiration4.xlsx")
    parser.add_argument("--out", type=Path, required=True, help="Output .xlsx path for the full Inspiration4 summary")
    parser.add_argument(
        "--overlap-out",
        type=Path,
        default=None,
        help="Optional output .xlsx path for the Inspiration4/Bed Rest overlap-only summary",
    )
    parser.add_argument(
        "--baseline-timepoints",
        nargs="+",
        default=DEFAULT_BASELINE_TIMEPOINTS,
        help="Baseline timepoints used for Baseline_ISV. Default: L-92 L-44 L-3",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    wide = read_csv_flex(args.wide)
    long_inventory = read_csv_flex(args.long)
    demographics = read_demographics(args.demographics)

    full_tables = build_summary_tables(
        wide=wide,
        long_inventory=long_inventory,
        demographics=demographics,
        baseline_timepoints=args.baseline_timepoints,
    )

    full_order = ["v3_Notes"] + BEDREST_ALIGNED_TABS + EXTRA_INSPIRATION4_TABS
    save_workbook(full_tables, args.out, ordered_names=full_order)

    print(f"Saved full Inspiration4 summary workbook to: {args.out}")

    if args.overlap_out is not None:
        overlap_tables = build_overlap_summary_tables(wide)
        save_workbook(
            overlap_tables,
            args.overlap_out,
            ordered_names=BEDREST_ALIGNED_TABS,
        )
        print(f"Saved Inspiration4/Bed Rest overlap summary workbook to: {args.overlap_out}")
        print(
            "Overlap variables:",
            overlap_tables["Dataset_Inventory"].loc[
                overlap_tables["Dataset_Inventory"]["Metric"] == "Canonical overlap variables",
                "Value",
            ].iloc[0],
        )


if __name__ == "__main__":
    main()
