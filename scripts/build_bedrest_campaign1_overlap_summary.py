#!/usr/bin/env python3
"""
build_bedrest_campaign1_overlap_summary.py

Create the LUNAR Bed Rest Campaign 1 overlap summary workbook from the raw Campaign 1
CSV files. The script:

1. Reads the raw Campaign 1 CSV files from either a directory or a ZIP archive.
2. Converts all files to a standardized long-format master table.
3. Harmonizes Campaign 1 variables to the Inspiration4 overlap variable names.
4. Adds three derived Inspiration4-compatible variables:
   - GLOBULIN = Total Protein - Albumin
   - ALBUMIN/GLOBULIN RATIO = Albumin / Globulin
   - BUN/CREATININE RATIO = BUN / blood-serum Creatinine
5. Writes a summary workbook with variable-level, timepoint-level, and subject-level QC tabs.

This script was written for LUNAR: Longitudinal Unification of Small-N Astronaut Responses.

Example:
    python build_bedrest_campaign1_overlap_summary.py \
        --input "data/raw/bed_rest_campaign1/Campaign 1-20260603T141236Z-3-001.zip" \
        --output-dir "outputs/bed_rest_campaign1"
"""

from __future__ import annotations

import argparse
import io
import math
import re
import zipfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DATASET_NAME = "Campaign 1"
OVERLAP_COHORT = ["C1G0001", "C1G0002", "C1G0003"]
TEST_PHASE_ORDER = {"PRE_TEST": 0, "IN_TEST": 1, "POST_TEST": 2}

# Raw variable aliases that map to Inspiration4-compatible variable names.
# This intentionally excludes urine creatinine and creatinine clearance from CREATININE.
RAW_TO_INSPIRATION4: Dict[str, str] = {
    # Albumin / protein
    "albumin": "ALBUMIN",
    "albumin g/dl": "ALBUMIN",
    "albumin g/dl": "ALBUMIN",
    "serum albumin": "ALBUMIN",
    "serum albumin g/dl": "ALBUMIN",
    "total protein": "PROTEIN; TOTAL",
    "total protein g/dl": "PROTEIN; TOTAL",
    # Chemistry
    "glucose": "GLUCOSE",
    "glucose mg/dl": "GLUCOSE",
    "blood urea nitrogen (bun)": "UREA NITROGEN (BUN)",
    "blood urea nitrogen (bun) mg/dl": "UREA NITROGEN (BUN)",
    "creatinine": "CREATININE",
    "creatinine mg/dl": "CREATININE",
    "serum creatinine": "CREATININE",
    "serum creatinine mg/dl": "CREATININE",
    "total bilirubin": "BILIRUBIN; TOTAL",
    "total bilirubin mg/dl": "BILIRUBIN; TOTAL",
    "bilirubin": "BILIRUBIN; TOTAL",
    "aspartate aminotransferase (ast)": "AST",
    "aspartate aminotransferase (ast) u/l": "AST",
    "ast": "AST",
    "ast u/l": "AST",
    "alanine aminotransferase (alt)": "ALT",
    "alanine aminotransferase (alt) u/l": "ALT",
    "alt": "ALT",
    "alt u/l": "ALT",
    "alkaline phosphatase (alp)": "ALKALINE PHOSPHATASE",
    "alkaline phosphatase (alp) u/l": "ALKALINE PHOSPHATASE",
    "alkphos": "ALKALINE PHOSPHATASE",
    "alkphos u/l": "ALKALINE PHOSPHATASE",
    "sodium": "SODIUM",
    "sodium mmol/l": "SODIUM",
    "pcba-sodium": "SODIUM",
    "pcba-sodium mmol/l": "SODIUM",
    "potassium": "POTASSIUM",
    "potassium mmol/l": "POTASSIUM",
    "pcba-potassium": "POTASSIUM",
    "pcba-potassium mmol/l": "POTASSIUM",
    "chloride": "CHLORIDE",
    "chloride mmol/l": "CHLORIDE",
    "calcium": "CALCIUM",
    "calcium mg/dl": "CALCIUM",
    "serum calcium (icp)": "CALCIUM",
    "serum calcium (icp) mg/dl": "CALCIUM",
    "serum calcium (icp) mmol/l": "CALCIUM",
    "pcba-ionized calcium": "CALCIUM",
    "pcba-ionized calcium mmol/l": "CALCIUM",
    "carbon dioxide": "CARBON DIOXIDE",
    "carbon dioxide mmol/l": "CARBON DIOXIDE",
    # Inflammation
    "c-reactive protein": "CRP",
    "c-reactive protein mg/dl": "CRP",
}

# Prevent false-positive matches to non-blood creatinine and urine/protein screening terms.
EXCLUDE_OVERLAP_PATTERNS = [
    re.compile(r"creatinine[, ]+urine", re.I),
    re.compile(r"urinary creatinine", re.I),
    re.compile(r"creatinine clearance", re.I),
    re.compile(r"n[- ]?telopeptide.*creatinine", re.I),
    re.compile(r"protein$", re.I),  # urine dipstick Protein should not become total protein
]

META_COLS = {
    "ID",
    "Subject",
    "Date",
    "Test_Phase",
    "BR_Day",
    "Panel",
    "Group",
    "GroupLabel",
    "A2_Status",
    "Accession Number",
    "Record Number",
    "Notes",
    "Session",
    "Statistic",
    "Time_Period",
    "Trial",
}

UNIT_PATTERNS = [
    r"mg/dL",
    r"g/dL",
    r"g/Dl",
    r"mmol/L",
    r"U/L",
    r"uIU/mL",
    r"ng/cL",
    r"pg/mL",
    r"mg/dl",
    r"mmol/l",
    r"nM/mM Cr",
    r"mg/day",
    r"mmol/day",
    r"umol/day",
    r"umol/d",
    r"uM/Day",
    r"mL",
    r"%",
]


def normalize_key(text: object) -> str:
    """Normalize strings for alias matching."""
    if pd.isna(text):
        return ""
    s = str(text).strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def parse_numeric_and_flag(value: object) -> Tuple[float, Optional[str]]:
    """Return first numeric token and any attached abnormal flag such as H/L/A.

    Examples:
        '6.4  (L)' -> (6.4, 'L')
        '5.30A  (H)' -> (5.30, 'A;H')
        'Negative' -> (nan, None)
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return np.nan, None
    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
        return float(value), None
    s = str(value).strip()
    if s == "" or s.lower() in {"nan", "na", "n/a"}:
        return np.nan, None
    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    num = float(match.group(0)) if match else np.nan
    flags: List[str] = []
    # Parenthetical flags e.g., (H), (L)
    flags.extend(re.findall(r"\(([A-Za-z]+)\)", s))
    # Letter directly following the numeric token e.g., 5.30A
    if match:
        tail = s[match.end() : match.end() + 2]
        direct = re.match(r"[A-Za-z]", tail)
        if direct:
            flags.append(direct.group(0))
    flags = [f.upper() for f in flags if f]
    return num, ";".join(sorted(set(flags))) if flags else None


def infer_unit(raw_var: str) -> Optional[str]:
    """Infer units from the column name when units are embedded in the header."""
    if raw_var is None:
        return None
    s = str(raw_var)
    for pat in UNIT_PATTERNS:
        if re.search(re.escape(pat), s, flags=re.I):
            # Preserve canonical capitalization for common units.
            found = re.search(re.escape(pat), s, flags=re.I).group(0)
            return found.replace("g/Dl", "g/dL").replace("mg/dl", "mg/dL").replace("mmol/l", "mmol/L")
    # Some variables encode units inside parentheses.
    m = re.search(r"\(([^)]*(?:%|mL|mg|g|mmol|U/L|pg|ng|uIU)[^)]*)\)", s, flags=re.I)
    if m:
        return m.group(1)
    return None


def canonicalize_variable(raw_var: str) -> str:
    """Remove common unit strings and punctuation for a cleaner canonical Campaign 1 name."""
    s = str(raw_var).strip()
    s = re.sub(r"\s+", " ", s)
    for pat in sorted(UNIT_PATTERNS, key=len, reverse=True):
        s = re.sub(re.escape(pat), "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -_/()")
    # Manual cleanups
    s = s.replace("PCBA-", "PCBA-").strip()
    return s


def infer_sample_type(source_file: str, raw_var: str) -> str:
    text = f"{source_file} {raw_var}".lower()
    if "urine" in text or "urinalysis" in text or "urinary" in text:
        return "Urine"
    if "breath" in text:
        return "Breath"
    if "salivary" in text:
        return "Saliva"
    if "plasma" in text:
        return "Plasma"
    if "serum" in text or "chem" in text or "blood" in text or "hematology" in text or "protein" in text or "bone_marker" in text or "pcba" in text or "hepatic_renal" in text or "creatinine_albumin" in text:
        return "Serum/Blood/Other"
    return "Other/Unspecified"


def map_to_inspiration4(source_file: str, raw_var: str, canonical: str, sample_type: str) -> Tuple[Optional[str], Optional[str], bool]:
    """Return matched Inspiration4 variable and match note."""
    rv = str(raw_var)
    for pat in EXCLUDE_OVERLAP_PATTERNS:
        if pat.search(rv):
            return None, None, False

    # Avoid mapping urine dipstick Protein to total serum protein.
    if sample_type == "Urine" and normalize_key(raw_var) == "protein":
        return None, None, False

    candidates = [normalize_key(raw_var), normalize_key(canonical)]
    for key in candidates:
        if key in RAW_TO_INSPIRATION4:
            return RAW_TO_INSPIRATION4[key], "Exact/related alias", True
    return None, None, False


def open_csv_files(input_path: Path) -> Iterable[Tuple[str, pd.DataFrame]]:
    """Yield (filename, dataframe) for all CSV files in a directory or ZIP archive."""
    if input_path.is_file() and input_path.suffix.lower() == ".zip":
        with zipfile.ZipFile(input_path, "r") as zf:
            for name in sorted(zf.namelist()):
                if name.lower().endswith(".csv") and not Path(name).name.startswith("."):
                    with zf.open(name) as f:
                        df = pd.read_csv(f)
                    yield Path(name).name, df
    elif input_path.is_dir():
        for csv_path in sorted(input_path.rglob("*.csv")):
            df = pd.read_csv(csv_path)
            yield csv_path.name, df
    else:
        raise FileNotFoundError(f"Input path must be a directory or .zip file: {input_path}")


def rows_from_standard_wide(source_file: str, df: pd.DataFrame) -> List[dict]:
    """Convert a standard wide assay file to long rows."""
    rows: List[dict] = []
    subject_col = "Subject" if "Subject" in df.columns else None
    value_cols = []
    for c in df.columns:
        if c in META_COLS or c.startswith("Unnamed"):
            continue
        # These are not observations.
        if c in {"Color", "Appearance"}:
            value_cols.append(c)  # Keep raw nonnumeric observations in master.
        else:
            value_cols.append(c)

    for _, rec in df.iterrows():
        subject = rec.get(subject_col, np.nan) if subject_col else np.nan
        group = rec.get("Group", rec.get("GroupLabel", np.nan))
        date = rec.get("Date", np.nan)
        test_phase = rec.get("Test_Phase", np.nan)
        br_day = rec.get("BR_Day", np.nan)
        panel = rec.get("Panel", np.nan)
        statistic = rec.get("Statistic", np.nan)
        time_period = rec.get("Time_Period", np.nan)
        trial = rec.get("Trial", np.nan)

        for raw_var in value_cols:
            raw_value = rec.get(raw_var, np.nan)
            if pd.isna(raw_value):
                # Keep explicit missingness only for overlap files would bloat the master;
                # skip fully blank cells from raw spreadsheets.
                continue
            value_num, value_flag = parse_numeric_and_flag(raw_value)
            unit = infer_unit(raw_var)
            canonical = canonicalize_variable(raw_var)
            sample_type = infer_sample_type(source_file, raw_var)
            i4_var, match_note, is_overlap = map_to_inspiration4(source_file, raw_var, canonical, sample_type)
            time_min = np.nan
            if normalize_key(raw_var) == "time (min)":
                canonical = "Time"
                unit = "min"
                time_min = value_num
            row = {
                "Dataset": DATASET_NAME,
                "Source_File": source_file,
                "Data_Level": "subject_observation",
                "Subject_ID": str(subject) if not pd.isna(subject) else np.nan,
                "Group": group,
                "Date": date,
                "Test_Phase": test_phase,
                "BR_Day": pd.to_numeric(br_day, errors="coerce"),
                "Panel": panel,
                "Variable_Raw": raw_var,
                "Variable_Canonical": canonical,
                "Value_Raw": raw_value,
                "Value_Numeric": value_num,
                "Value_Flag": value_flag,
                "Unit": unit,
                "Statistic": statistic,
                "Time_Period": time_period,
                "Trial": trial,
                "Time_Min": time_min,
                "Match_To_Inspiration4": match_note,
                "Has_Numeric_Value": bool(not pd.isna(value_num)),
                "Has_Raw_Value": True,
                "Inspiration4_Variable": i4_var,
                "Is_Inspiration4_Overlap": bool(is_overlap),
                "Sample_Type": sample_type,
            }
            rows.append(row)
    return rows


def rows_from_long_value_file(source_file: str, df: pd.DataFrame) -> List[dict]:
    """Handle files where variable names are in Unit and observations are in Value."""
    rows: List[dict] = []
    for _, rec in df.iterrows():
        raw_var = rec.get("Unit", np.nan)
        raw_value = rec.get("Value", np.nan)
        if pd.isna(raw_var) or pd.isna(raw_value):
            continue
        value_num, value_flag = parse_numeric_and_flag(raw_value)
        canonical = canonicalize_variable(str(raw_var))
        sample_type = infer_sample_type(source_file, str(raw_var))
        i4_var, match_note, is_overlap = map_to_inspiration4(source_file, str(raw_var), canonical, sample_type)
        rows.append(
            {
                "Dataset": DATASET_NAME,
                "Source_File": source_file,
                "Data_Level": "group_summary" if "avgs" in source_file.lower() else "subject_observation",
                "Subject_ID": str(rec.get("Subject", np.nan)) if "Subject" in rec.index and not pd.isna(rec.get("Subject")) else np.nan,
                "Group": rec.get("GroupLabel", rec.get("Group", np.nan)),
                "Date": rec.get("Date", np.nan),
                "Test_Phase": rec.get("Test_Phase", np.nan),
                "BR_Day": pd.to_numeric(rec.get("BR_Day", np.nan), errors="coerce"),
                "Panel": np.nan,
                "Variable_Raw": str(raw_var),
                "Variable_Canonical": canonical,
                "Value_Raw": raw_value,
                "Value_Numeric": value_num,
                "Value_Flag": value_flag,
                "Unit": infer_unit(str(raw_var)),
                "Statistic": rec.get("Statistic", np.nan),
                "Time_Period": rec.get("Time_Period", np.nan),
                "Trial": rec.get("Trial", np.nan),
                "Time_Min": np.nan,
                "Match_To_Inspiration4": match_note,
                "Has_Numeric_Value": bool(not pd.isna(value_num)),
                "Has_Raw_Value": True,
                "Inspiration4_Variable": i4_var,
                "Is_Inspiration4_Overlap": bool(is_overlap),
                "Sample_Type": sample_type,
            }
        )
    return rows


def build_master_long(input_path: Path) -> pd.DataFrame:
    """Build a long-format master table from all raw CSV files."""
    all_rows: List[dict] = []
    for source_file, df in open_csv_files(input_path):
        if {"Unit", "Value"}.issubset(df.columns):
            all_rows.extend(rows_from_long_value_file(source_file, df))
        else:
            all_rows.extend(rows_from_standard_wide(source_file, df))
    master = pd.DataFrame(all_rows)
    if master.empty:
        raise ValueError("No long-format rows were produced. Check input files.")
    # Normalize subject IDs and test phase text.
    master["Subject_ID"] = master["Subject_ID"].astype(str).str.strip()
    master.loc[master["Subject_ID"].isin(["nan", "None", ""]), "Subject_ID"] = np.nan
    master["Test_Phase"] = master["Test_Phase"].astype(str).str.strip()
    master.loc[master["Test_Phase"].isin(["nan", "None", ""]), "Test_Phase"] = np.nan
    return master


def derive_overlap_variables(overlap: pd.DataFrame) -> pd.DataFrame:
    """Append derived variables to the measured overlap table."""
    measured = overlap.copy()
    measured["Derivation_Method"] = "Measured"
    for col in ["Derived_From", "Calculation_Formula", "Component_Values"]:
        if col not in measured.columns:
            measured[col] = np.nan

    # Only derive from true C1G overlap cohort and blood/serum/non-urine rows.
    chem = measured[
        measured["Subject_ID"].isin(OVERLAP_COHORT)
        & measured["Has_Numeric_Value"].fillna(False).astype(bool)
        & ~measured["Sample_Type"].astype(str).str.contains("Urine", case=False, na=False)
    ].copy()

    derived_rows: List[dict] = []

    def add_row(base: pd.Series, var: str, value: float, unit: str, source: str, formula: str, components: str) -> None:
        row = {c: np.nan for c in measured.columns}
        for c in ["Dataset", "Subject_ID", "Group", "Date", "Test_Phase", "BR_Day"]:
            row[c] = base.get(c, np.nan)
        row.update(
            {
                "Source_File": source,
                "Data_Level": "subject_observation",
                "Panel": "Derived chemistry",
                "Variable_Raw": var,
                "Variable_Canonical": var,
                "Value_Raw": round(float(value), 6),
                "Value_Numeric": float(value),
                "Value_Flag": np.nan,
                "Unit": unit,
                "Time_Min": np.nan,
                "Match_To_Inspiration4": "Derived from overlapping Campaign 1 blood chemistry variables",
                "Has_Numeric_Value": True,
                "Has_Raw_Value": True,
                "Inspiration4_Variable": var,
                "Is_Inspiration4_Overlap": True,
                "Sample_Type": "Serum/Blood/Other",
                "Derived_From": source,
                "Calculation_Formula": formula,
                "Component_Values": components,
                "Derivation_Method": "Derived",
            }
        )
        derived_rows.append(row)

    # Protein-derived variables from MR016G Protein file: Albumin + Total Protein same file.
    protein_src = "MR016G_Campaign_1_Nutritional_Status_Assessment_Protein.csv"
    prot = chem[chem["Source_File"].eq(protein_src) & chem["Inspiration4_Variable"].isin(["ALBUMIN", "PROTEIN; TOTAL"])]
    if not prot.empty:
        piv = prot.pivot_table(
            index=["Dataset", "Subject_ID", "Date", "Test_Phase", "BR_Day"],
            columns="Inspiration4_Variable",
            values="Value_Numeric",
            aggfunc="first",
        ).reset_index()
        for _, r in piv.dropna(subset=["ALBUMIN", "PROTEIN; TOTAL"]).iterrows():
            glob = r["PROTEIN; TOTAL"] - r["ALBUMIN"]
            comps = f"Total Protein={r['PROTEIN; TOTAL']} g/dL; Albumin={r['ALBUMIN']} g/dL"
            add_row(r, "GLOBULIN", glob, "g/dL", f"DERIVED_FROM_{protein_src}", "Globulin = Total Protein - Albumin", comps)
            if glob != 0:
                add_row(r, "ALBUMIN/GLOBULIN RATIO", r["ALBUMIN"] / glob, "ratio", f"DERIVED_FROM_{protein_src}", "Albumin/Globulin Ratio = Albumin / (Total Protein - Albumin)", comps)

    # Protein-derived variables from MR010G Flight Chem total protein + MR010G Other albumin.
    fc_src = "MR010G_Campaign_1_Flight_Chem_Profile.csv"
    other_src = "MR010G_Campaign_1_Other.csv"
    tp = chem[chem["Source_File"].eq(fc_src) & chem["Inspiration4_Variable"].eq("PROTEIN; TOTAL")]
    alb = chem[chem["Source_File"].eq(other_src) & chem["Inspiration4_Variable"].eq("ALBUMIN")]
    if not tp.empty and not alb.empty:
        key_cols = ["Dataset", "Subject_ID", "Date", "Test_Phase", "BR_Day"]
        tp2 = tp[key_cols + ["Value_Numeric"]].rename(columns={"Value_Numeric": "PROTEIN; TOTAL"})
        alb2 = alb[key_cols + ["Value_Numeric"]].rename(columns={"Value_Numeric": "ALBUMIN"})
        m = pd.merge(tp2, alb2, on=key_cols, how="inner").dropna(subset=["ALBUMIN", "PROTEIN; TOTAL"])
        for _, r in m.iterrows():
            glob = r["PROTEIN; TOTAL"] - r["ALBUMIN"]
            comps = f"Total Protein={r['PROTEIN; TOTAL']} g/dL; Albumin={r['ALBUMIN']} g/dL"
            source = "DERIVED_FROM_MR010G_Flight_Chem_Profile + MR010G_Other"
            add_row(r, "GLOBULIN", glob, "g/dL", source, "Globulin = Total Protein - Albumin", comps)
            if glob != 0:
                add_row(r, "ALBUMIN/GLOBULIN RATIO", r["ALBUMIN"] / glob, "ratio", source, "Albumin/Globulin Ratio = Albumin / (Total Protein - Albumin)", comps)

    # BUN/Creatinine ratio: use only same-source MR010G Flight Chem blood/serum creatinine.
    fc = chem[chem["Source_File"].eq(fc_src) & chem["Inspiration4_Variable"].isin(["UREA NITROGEN (BUN)", "CREATININE"])]
    if not fc.empty:
        piv = fc.pivot_table(
            index=["Dataset", "Subject_ID", "Date", "Test_Phase", "BR_Day"],
            columns="Inspiration4_Variable",
            values="Value_Numeric",
            aggfunc="first",
        ).reset_index()
        for _, r in piv.dropna(subset=["UREA NITROGEN (BUN)", "CREATININE"]).iterrows():
            if r["CREATININE"] != 0:
                comps = f"BUN={r['UREA NITROGEN (BUN)']} mg/dL; blood/serum Creatinine={r['CREATININE']} mg/dL"
                add_row(
                    r,
                    "BUN/CREATININE RATIO",
                    r["UREA NITROGEN (BUN)"] / r["CREATININE"],
                    "ratio",
                    "DERIVED_FROM_MR010G_Campaign_1_Flight_Chem_Profile_serum_blood_only",
                    "BUN/Creatinine Ratio = Blood Urea Nitrogen / blood-serum Creatinine",
                    comps,
                )

    derived = pd.DataFrame(derived_rows, columns=measured.columns)
    combined = pd.concat([measured, derived], ignore_index=True)
    combined["Has_Numeric_Value"] = combined["Value_Numeric"].notna()
    combined["Has_Raw_Value"] = combined["Value_Raw"].notna()
    combined["Is_Inspiration4_Overlap"] = True
    return combined


def quantile_25(s: pd.Series) -> float:
    return s.quantile(0.25)


def quantile_75(s: pd.Series) -> float:
    return s.quantile(0.75)


def stats_table(df: pd.DataFrame, group_cols: List[str]) -> pd.DataFrame:
    """Summary statistics for numeric values, grouped by requested columns."""
    num = df.dropna(subset=["Value_Numeric"]).copy()
    if num.empty:
        return pd.DataFrame(columns=group_cols + ["N", "Mean", "SD", "Median", "Min", "Q1", "Q3", "IQR", "Max"])
    out = (
        num.groupby(group_cols, dropna=False)["Value_Numeric"]
        .agg(N="count", Mean="mean", SD="std", Median="median", Min="min", Q1=quantile_25, Q3=quantile_75, Max="max")
        .reset_index()
    )
    out["IQR"] = out["Q3"] - out["Q1"]
    out = out[group_cols + ["N", "Mean", "SD", "Median", "Min", "Q1", "Q3", "IQR", "Max"]]
    return out.sort_values(group_cols).reset_index(drop=True)


def make_missingness_tables(overlap: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Create missingness using subject x phase x variable coverage, not raw row counts."""
    subjects = OVERLAP_COHORT
    phases = ["PRE_TEST", "IN_TEST", "POST_TEST"]
    variables = sorted(overlap["Inspiration4_Variable"].dropna().unique())

    present = (
        overlap[overlap["Subject_ID"].isin(subjects) & overlap["Test_Phase"].isin(phases)]
        .groupby(["Subject_ID", "Test_Phase", "Inspiration4_Variable"])["Value_Numeric"]
        .apply(lambda s: s.notna().any())
        .reset_index(name="Present")
    )
    grid = pd.MultiIndex.from_product([subjects, phases, variables], names=["Subject_ID", "Test_Phase", "Inspiration4_Variable"]).to_frame(index=False)
    cov = pd.merge(grid, present, on=["Subject_ID", "Test_Phase", "Inspiration4_Variable"], how="left")
    cov["Present"] = cov["Present"].fillna(False)

    missing_by_variable = cov.groupby("Inspiration4_Variable").agg(Expected=("Present", "size"), Present=("Present", "sum")).reset_index()
    missing_by_variable["Missing"] = missing_by_variable["Expected"] - missing_by_variable["Present"]
    missing_by_variable["Percent_Missing"] = missing_by_variable["Missing"] / missing_by_variable["Expected"] * 100

    subject_missing = cov.groupby("Subject_ID").agg(Expected=("Present", "size"), Present=("Present", "sum")).reset_index()
    subject_missing["Missing"] = subject_missing["Expected"] - subject_missing["Present"]
    subject_missing["Percent_Missing"] = subject_missing["Missing"] / subject_missing["Expected"] * 100

    subject_time = cov.groupby(["Subject_ID", "Test_Phase"]).agg(Expected_Variables=("Present", "size"), Variables_Present=("Present", "sum")).reset_index()
    subject_time["Variables_Missing"] = subject_time["Expected_Variables"] - subject_time["Variables_Present"]
    subject_time["Percent_Missing"] = subject_time["Variables_Missing"] / subject_time["Expected_Variables"] * 100
    missing_names = cov[~cov["Present"]].groupby(["Subject_ID", "Test_Phase"])["Inspiration4_Variable"].apply(lambda x: "; ".join(sorted(x))).reset_index(name="Missing_Variables")
    subject_time = pd.merge(subject_time, missing_names, on=["Subject_ID", "Test_Phase"], how="left")
    subject_time["Test_Phase_Order"] = subject_time["Test_Phase"].map(TEST_PHASE_ORDER)
    subject_time = subject_time.sort_values(["Subject_ID", "Test_Phase_Order"]).drop(columns="Test_Phase_Order").reset_index(drop=True)

    return missing_by_variable, subject_missing, subject_time


def build_summary_tables(master: pd.DataFrame, overlap: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Build all workbook tabs."""
    overlap = overlap[overlap["Subject_ID"].isin(OVERLAP_COHORT)].copy()
    variables = sorted(overlap["Inspiration4_Variable"].dropna().unique())
    measured_variables = sorted(overlap.loc[overlap["Derivation_Method"].eq("Measured"), "Inspiration4_Variable"].dropna().unique())
    derived_variables = sorted(overlap.loc[overlap["Derivation_Method"].eq("Derived"), "Inspiration4_Variable"].dropna().unique())

    dataset_inventory = pd.DataFrame(
        [
            ["Dataset", "Campaign 1 Bed Rest / LUNAR Inspiration4 overlap subset"],
            ["Overlap cohort participants", len(OVERLAP_COHORT)],
            ["Participant IDs", ", ".join(OVERLAP_COHORT)],
            ["Test phases", ", ".join(["PRE_TEST", "IN_TEST", "POST_TEST"])],
            ["Measured overlap variables", len(measured_variables)],
            ["Derived variables", len(derived_variables)],
            ["Total overlap variables", len(variables)],
            ["Overlap rows", len(overlap)],
            ["Measured overlap rows", int(overlap["Derivation_Method"].eq("Measured").sum())],
            ["Derived rows", int(overlap["Derivation_Method"].eq("Derived").sum())],
            ["Raw master rows generated", len(master)],
            ["Raw subject IDs in all Campaign 1 files", master["Subject_ID"].dropna().nunique()],
            ["Creatinine QC", "BUN/CREATININE RATIO uses only MR010G Flight Chemistry blood/serum creatinine; urine creatinine and creatinine clearance are excluded."],
            ["Protein derivation QC", "Globulin and A/G ratio are derived only from paired total protein + albumin within the same subject/date/timepoint."],
            ["Generated by", "build_bedrest_campaign1_overlap_summary.py"],
        ],
        columns=["Metric", "Value"],
    )

    variable_stats = stats_table(overlap, ["Inspiration4_Variable"])
    variable_stats_by_timepoint = stats_table(overlap, ["Inspiration4_Variable", "Test_Phase"])
    subject_stats = stats_table(overlap, ["Subject_ID", "Inspiration4_Variable"])
    subject_timepoint_variable_stats = stats_table(overlap, ["Subject_ID", "Test_Phase", "Inspiration4_Variable"])

    missing_by_variable, subject_missingness, subject_timepoint_stats = make_missingness_tables(overlap)
    missing_by_timepoint = subject_timepoint_stats.groupby("Test_Phase").agg(Expected=("Expected_Variables", "sum"), Present=("Variables_Present", "sum")).reset_index()
    missing_by_timepoint["Missing"] = missing_by_timepoint["Expected"] - missing_by_timepoint["Present"]
    missing_by_timepoint["Percent_Missing"] = missing_by_timepoint["Missing"] / missing_by_timepoint["Expected"] * 100
    missing_by_timepoint["Test_Phase_Order"] = missing_by_timepoint["Test_Phase"].map(TEST_PHASE_ORDER)
    missing_by_timepoint = missing_by_timepoint.sort_values("Test_Phase_Order").drop(columns="Test_Phase_Order")

    crosswalk = pd.DataFrame(
        {
            "Inspiration4_Variable": variables,
            "Campaign1_Variable_Canonical": variables,
            "Match_Type": ["Derived from Campaign 1 blood chemistry" if v in derived_variables else "Measured exact/alias overlap" for v in variables],
        }
    )

    units_audit = (
        overlap.groupby("Inspiration4_Variable", dropna=False)
        .agg(
            Units_Observed=("Unit", lambda s: "; ".join(sorted(set(str(x) for x in s.dropna() if str(x) != "nan")))),
            Unit_Count=("Unit", lambda s: len(set(str(x) for x in s.dropna() if str(x) != "nan"))),
            Rows=("Unit", "size"),
        )
        .reset_index()
    )
    units_audit["Unit_Consistency_Flag"] = np.where(units_audit["Unit_Count"].le(1), "Single unit observed or blank", "Multiple units observed - review before cross-dataset comparison")

    derived = overlap[overlap["Derivation_Method"].eq("Derived")].copy()
    derived_audit = (
        derived.groupby("Inspiration4_Variable", dropna=False)
        .agg(
            Derived_Rows=("Inspiration4_Variable", "size"),
            Formula=("Calculation_Formula", lambda s: "; ".join(sorted(set(str(x) for x in s.dropna())))),
            Source_Files=("Derived_From", lambda s: "; ".join(sorted(set(str(x) for x in s.dropna())))),
            Example_Component_Values=("Component_Values", lambda s: next((str(x) for x in s.dropna()), "")),
        )
        .reset_index()
        if not derived.empty
        else pd.DataFrame(columns=["Inspiration4_Variable", "Derived_Rows", "Formula", "Source_Files", "Example_Component_Values"])
    )

    raw_subject_counts = master.groupby("Subject_ID", dropna=False).size().reset_index(name="Raw_Master_Row_Count")
    raw_subject_counts = raw_subject_counts.rename(columns={"Subject_ID": "Raw_Subject_ID"})
    raw_subject_counts["Looks_Like_Campaign1_Participant"] = raw_subject_counts["Raw_Subject_ID"].isin(OVERLAP_COHORT)
    raw_subject_counts["Included_In_Overlap_Cohort"] = raw_subject_counts["Raw_Subject_ID"].isin(OVERLAP_COHORT)
    raw_subject_counts["Notes"] = np.where(
        raw_subject_counts["Included_In_Overlap_Cohort"],
        "Included C1G participant in overlap cohort",
        "Raw/non-overlap subject/sample/group identifier; not part of clinical chemistry overlap cohort",
    )
    participant_audit = raw_subject_counts.sort_values(["Included_In_Overlap_Cohort", "Raw_Subject_ID"], ascending=[False, True]).reset_index(drop=True)

    # Keep a clean overlap data tab.
    overlap_out = overlap.sort_values(["Subject_ID", "Test_Phase", "BR_Day", "Date", "Inspiration4_Variable", "Derivation_Method"], na_position="last").reset_index(drop=True)

    return {
        "Dataset_Inventory": dataset_inventory,
        "Variable_Stats": variable_stats,
        "Variable_Stats_By_Timepoint": variable_stats_by_timepoint,
        "Subject_Stats": subject_stats,
        "Subject_Timepoint_Variable_Stat": subject_timepoint_variable_stats,
        "Subject_Missingness": subject_missingness,
        "Subject_Timepoint_Stats": subject_timepoint_stats,
        "Missing_By_Variable": missing_by_variable,
        "Missing_By_Timepoint": missing_by_timepoint,
        "Overlap_Crosswalk": crosswalk,
        "Units_Audit": units_audit,
        "Derived_Variables_Audit": derived_audit,
        "Participant_Audit": participant_audit,
        "Overlap_Data": overlap_out,
    }


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9D9D9")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        # Header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Body style
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(bottom=thin_gray)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    if "Percent" in str(ws.cell(row=1, column=cell.column).value):
                        cell.number_format = "0.0"
                    else:
                        cell.number_format = "0.000"
        # Widths
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            vals = [ws.cell(row=r, column=col_idx).value for r in range(1, min(ws.max_row, 100) + 1)]
            max_len = max([len(str(v)) if v is not None else 0 for v in vals] + [8])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)
        # Excel table styling where possible.
        if ws.max_row >= 2 and ws.max_column >= 1:
            safe_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:25]
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tab = Table(displayName=f"tbl_{safe_name}", ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            try:
                ws.add_table(tab)
            except Exception:
                pass
    wb.save(path)


def write_outputs(tables: Dict[str, pd.DataFrame], master: pd.DataFrame, output_dir: Path, prefix: str) -> Tuple[Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / f"{prefix}_Summary.xlsx"
    overlap_csv_path = output_dir / f"{prefix}_Overlap_Data.csv"
    master_csv_path = output_dir / f"{prefix}_Master_Long.csv"

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet, table in tables.items():
            # Excel sheet names must be <=31 characters.
            safe_sheet = sheet[:31]
            table.to_excel(writer, sheet_name=safe_sheet, index=False)

    style_workbook(workbook_path)
    tables["Overlap_Data"].to_csv(overlap_csv_path, index=False)
    master.to_csv(master_csv_path, index=False)
    return workbook_path, overlap_csv_path, master_csv_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the LUNAR Bed Rest Campaign 1 overlap summary workbook from raw CSV files.")
    parser.add_argument("--input", required=True, help="Path to raw Campaign 1 CSV folder or ZIP file.")
    parser.add_argument("--output-dir", default="outputs/bed_rest_campaign1", help="Directory for output workbook and CSV files.")
    parser.add_argument("--prefix", default="LUNAR_BedRest_Campaign1_OVERLAP_v3", help="Output filename prefix.")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)

    master = build_master_long(input_path)
    measured_overlap = master[master["Is_Inspiration4_Overlap"].fillna(False)].copy()
    measured_overlap = measured_overlap[measured_overlap["Subject_ID"].isin(OVERLAP_COHORT)].copy()
    overlap = derive_overlap_variables(measured_overlap)
    tables = build_summary_tables(master, overlap)
    workbook_path, overlap_csv_path, master_csv_path = write_outputs(tables, master, output_dir, args.prefix)

    print("Created:")
    print(f"  Summary workbook: {workbook_path}")
    print(f"  Overlap data CSV: {overlap_csv_path}")
    print(f"  Master long CSV: {master_csv_path}")
    print("QC:")
    print(f"  Raw master rows: {len(master)}")
    print(f"  Overlap rows: {len(tables['Overlap_Data'])}")
    print(f"  Overlap variables: {tables['Overlap_Data']['Inspiration4_Variable'].nunique()}")
    print(f"  Overlap cohort participants: {', '.join(OVERLAP_COHORT)}")


if __name__ == "__main__":
    main()
